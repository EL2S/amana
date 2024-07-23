from datetime import datetime
from django.shortcuts import render,redirect, get_object_or_404
from django.contrib.auth import logout, authenticate, login as amana_login
from django.contrib.auth.decorators import login_required, permission_required
from django.utils import timezone
from django.contrib import messages
from django.contrib.auth.models import User, Group, Permission
from django.core.paginator import Paginator
from django.contrib.contenttypes.models import ContentType
from django.db.models import Q, F, Sum, Case, When, IntegerField, ExpressionWrapper, DecimalField
from django.contrib import messages
from django.utils import timezone
from amana.models import Car, Disbursement, Health, Insurance, Insured, InvoiceHealth, Journey, PrimeAssurance, Provider, Sex, SexUser, SmallCrate
from decimal import Decimal
import locale
from django.http import HttpResponse, JsonResponse
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from django.db import transaction
import json
from django.db.models.functions import Cast
from django.db.models.functions import TruncDate
from django.db.models import DateField
import random
import os
from django.conf import settings
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from dateutil.relativedelta import relativedelta
import re

def error(request, exception):
    return render(request, 'error.html',status=403)

def user_login(request):
    context = {}
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            user.last_login = timezone.now()
            user.save()
            amana_login(request, user)
            if user.is_superuser:
                return redirect('view_user')
            else:
                user_permission_ids = user.user_permissions.values_list('id', flat=True)
                permissions = Permission.objects.filter(id__in=user_permission_ids, codename__startswith='view')
                group_permissions = Permission.objects.filter(group__user=user,codename__startswith='view').distinct()
                if permissions.exists():
                    perm = random.choice(permissions)
                    return redirect(perm.codename)
                elif group_permissions.exists():
                        perm = random.choice(group_permissions)
                        return redirect(perm.codename)
                else:
                    context['error_message'] = 'Aucune permission de vue trouvée.'
                    return render(request, 'login.html', context)
        else:
            context['error_message'] = 'Nom d\'utilisateur ou mot de passe incorrect'
            return render(request, 'login.html', context)
    return render(request, 'login.html',context)

@permission_required('amana.view_dashboard', raise_exception=True)
def home(request):
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    
    context ={
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser
    }
    return render(request, 'dashboard.html',context)

@login_required
def user_logout(request):
    logout(request)
    return redirect('login')

@permission_required('auth.delete_user', raise_exception=True)
def delete_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.delete()
    return redirect('view_user')

@login_required
def change_password(request, user_id):
    context = {}
    user = get_object_or_404(User, id=user_id)
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context.update({
        'user_change': user,
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
        'user_id':user_id
    })
    if request.method == 'POST':
        password = request.POST['password']
        confirm_password = request.POST['confirm_password']
        if password != confirm_password:
            context['error_message'] = 'Les mots de passe ne correspondent pas.'
            return render(request, 'password.html', context)
        user.set_password(password)
        user.save()
        return redirect('change_user', user_id=user_id)
    return render(request, 'password.html', context)

@permission_required('auth.view_user', raise_exception=True)
def view_user(request):
    search_query = request.POST.get('search', '')
    users = User.objects.all()
    if search_query:
        users = users.filter(Q(first_name__icontains=search_query) | Q(last_name__icontains=search_query) | Q(username__icontains=search_query) | Q(email__icontains=search_query))
    users = users.order_by('-id')
    paginator = Paginator(users, 12)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'users': page_obj,
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
    }
    return render(request, 'user/view_user.html', context)

@permission_required('auth.add_user', raise_exception=True)
def add_user(request):
    context = {}
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    sexes = Sex.objects.all()
    context.update({
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
        'sexes': sexes
    })

    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        sex_id = request.POST['sex']
        confirm_password = request.POST['confirm_password']

        if password != confirm_password:
            context['error_message'] = 'Les mots de passe ne correspondent pas.'
            return render(request, 'user/add_user.html', context)

        if User.objects.filter(username=username).exists():
            context['error_message'] = 'Le nom d\'utilisateur existe déjà.'
            return render(request, 'user/add_user.html', context)

        add_user = User.objects.create_user(username=username, password=password)
        add_user.is_active = True
        add_user.is_staff = True
        add_user.save()

        sex = SexUser.objects.create(sex_id=sex_id, user_id=add_user.id)
        sex.save()

        return redirect('change_user', add_user.id)

    return render(request, 'user/add_user.html', context)

@permission_required('auth.change_user', raise_exception=True)
def change_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user_sex = SexUser.objects.filter(user=user_id).first()

    amana_and_auth_content_types = ContentType.objects.filter(
        Q(app_label='amana') | Q(app_label='auth', model='user') | Q(app_label='admin', model='logentry')
    ).exclude(model__in=['sex', 'sexuser','disbursement'])

    user_permissions = user.user_permissions.filter(content_type__in=amana_and_auth_content_types)

    unassigned_permissions = Permission.objects.filter(content_type__in=amana_and_auth_content_types).exclude(
    id__in=user_permissions.values_list('id', flat=True)
    ).exclude(
        Q(content_type__app_label='amana', content_type__model='summary', codename__in=['add_summary', 'change_summary','delete_summary']) | Q(content_type__app_label='amana', content_type__model='production', codename__in=['add_production', 'change_production','delete_production']) | Q(content_type__app_label='admin', content_type__model='logentry', codename__in=['add_logentry', 'change_logentry']) | Q(content_type__app_label='amana', content_type__model='dashboard', codename__in=['add_dashboard', 'change_dashboard','delete_dashboard'])
    )
    all_groups = Group.objects.all()
    user_groups = user.groups.all()
    unassigned_groups = all_groups.difference(user_groups)
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    current_user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = current_user_sex.sex.sex_name.capitalize() if current_user_sex else 'Not specified'
    sexes = Sex.objects.all()
    context = {
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
        'sexes':sexes,
        'user_permissions': user_permissions,
        'groups': unassigned_groups,
        'user_groups': user_groups,
        'permissions': unassigned_permissions,
        'user_change': user,
        'user_sex': user_sex,
    }
    if request.method == 'POST':
        username = request.POST['username']
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        email = request.POST['email']
        is_active = request.POST.get('is_active', False) == 'on'
        is_superuser = request.POST.get('is_superuser', False) == 'on'
        sex = request.POST['sex']
        if user.username != username:
            if User.objects.filter(username=username).exists():
                context['error_message'] = 'Le nom d\'utilisateur existe déjà.'
                return render(request, 'user/change_user.html', context)
        
        if user.email != email:
            if User.objects.filter(email=email).exists():
                context['error_message'] = 'L\'adresse e-mail existe déjà.'
                return render(request, 'user/change_user.html', context)
            
        user_sex.sex_id = sex
        user_sex.save()
        user.username = username
        user.first_name = first_name
        user.last_name = last_name
        user.email = email
        user.is_active = is_active
        user.is_superuser = is_superuser
        user.save()
        aut_rem_options = request.POST.getlist('aut_rem')
        for permission_id in aut_rem_options:
            permission = Permission.objects.get(id=permission_id)
            if permission in unassigned_permissions:
                user.user_permissions.add(permission)
        for permission in user_permissions:
            if str(permission.id) not in aut_rem_options:
                user.user_permissions.remove(permission)
        group_rem_options = request.POST.getlist('group_rem')
        for group_id in group_rem_options:
            group = Group.objects.get(id=group_id)
            if group in unassigned_groups:
                user.groups.add(group)
        for group in user_groups:
            if str(group.id) not in group_rem_options:
                user.groups.remove(group)
        return redirect('view_user')
    return render(request, "user/change_user.html", context)

@permission_required('auth.view_group', raise_exception=True)
def view_group(request):
    search_query = request.POST.get('search', '')
    groups = Group.objects.all()
    if search_query:  
        groups = groups.filter(name__icontains=search_query)

    groups = groups.order_by('-id')
    paginator = Paginator(groups, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
       'groups': page_obj,
    }
    return render(request, 'group/view_group.html', context)

@permission_required('auth.change_group', raise_exception=True)
def change_group(request, group_id):
    group = get_object_or_404(Group, id=group_id)

    amana_and_auth_content_types = ContentType.objects.filter(
        Q(app_label='amana') | Q(app_label='auth', model='user') | Q(app_label='admin', model='logentry')
    ).exclude(model__in=['sex', 'sexuser','disbursement'])

    group_permissions = group.permissions.filter(content_type__in=amana_and_auth_content_types)

    unassigned_permissions = Permission.objects.filter(content_type__in=amana_and_auth_content_types).exclude(
    id__in=group_permissions.values_list('id', flat=True)
    ).exclude(
        Q(content_type__app_label='amana', content_type__model='summary', codename__in=['add_summary', 'change_summary','delete_summary']) | Q(content_type__app_label='amana', content_type__model='production', codename__in=['add_production', 'change_production','delete_production']) | Q(content_type__app_label='admin', content_type__model='logentry', codename__in=['add_logentry', 'change_logentry']) | Q(content_type__app_label='amana', content_type__model='dashboard', codename__in=['add_dashboard', 'change_dashboard','delete_dashboard'])
    )
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
        'group': group, 
        'all_permissions': unassigned_permissions, 
        'group_permissions': group_permissions
    }
    if request.method == 'POST':
        name = request.POST['name']
        if group.name != name:
            if Group.objects.filter(name=name).exists():
                context['error_message'] = 'Le nom du groupe existe déjà.'
                return render(request, 'group/change_group.html', context)
        group.name = name
        group.save()
        aut_rem_options = request.POST.getlist('aut_rem')
        
        # Ajouter de nouvelles autorisations sélectionnées
        for permission_id in aut_rem_options:
            permission = Permission.objects.get(id=permission_id)
            if permission in unassigned_permissions:
                group.permissions.add(permission)

        # Supprimer les autorisations désélectionnées
        for permission in group_permissions:
            if str(permission.id) not in aut_rem_options:
                group.permissions.remove(permission)
        return redirect('view_group')
    
    return render(request, 'group/change_group.html', context)

@permission_required('auth.delete_group', raise_exception=True)
def delete_group(request, group_id):
    group = get_object_or_404(Group, id=group_id)
    group.permissions.clear()
    group.delete()
    return redirect('view_group')

@permission_required('auth.add_group', raise_exception=True)
def add_group(request):
    amana_and_auth_content_types = ContentType.objects.filter(
        Q(app_label='amana') | Q(app_label='auth', model='user') | Q(app_label='admin', model='logentry')
    ).exclude(model__in=['sex', 'sexuser','disbursement'])

    permissions = Permission.objects.filter(content_type__in=amana_and_auth_content_types).exclude(
        Q(content_type__app_label='amana', content_type__model='summary', codename__in=['add_summary', 'change_summary','delete_summary']) | Q(content_type__app_label='amana', content_type__model='production', codename__in=['add_production', 'change_production','delete_production']) | Q(content_type__app_label='admin', content_type__model='logentry', codename__in=['add_logentry', 'change_logentry']) | Q(content_type__app_label='amana', content_type__model='dashboard', codename__in=['add_dashboard', 'change_dashboard','delete_dashboard'])
    )
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
        'permissions': permissions,
    }
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            if Group.objects.filter(name=name).exists():
                context['error_message'] = 'Le nom du groupe existe déjà.'
                return render(request, 'group/add_group.html', context)
            new_group = Group.objects.create(name=name)
            aut_rem_options = request.POST.getlist('aut_rem')
            for permission_id in aut_rem_options:
                new_group.permissions.add(permission_id)
            return redirect('view_group')
    return render(request, 'group/add_group.html', context)

@permission_required('amana.view_insurance', raise_exception=True)
def view_insurance(request):
    search_query = request.POST.get('search', '')
    insurances = Insurance.objects.all()
    if search_query:
        insurances = insurances.filter(Q(policy_number__icontains=search_query) | Q(certificate_number__icontains=search_query))

    insurances = insurances.order_by('-id')
    paginator = Paginator(insurances, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
       'insurances': page_obj,
       'search_query':search_query
    }
    return render(request, 'insurance/view_insurance.html', context)

@permission_required('amana.add_insurance', raise_exception=True)
def add_insurance(request):
    current_user = request.user
    first_name_current = current_user.first_name.capitalize()
    last_name_current = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name_current[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'first_name': first_name_current,
        'last_name': last_name_current,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
    }
    if request.method == 'POST':
        certificate_number = request.POST.get('certificate_number')
        policy_number = request.POST.get('policy_number')
        insured = request.POST.get('insured')
        period = request.POST.get('period')
        address = request.POST.get('address')
        phone = request.POST.get('phone')
        insurance_type = request.POST.get('insurance_type')
        phone = request.POST.get('phone')
        deposit = request.POST.get('deposit')
        balance = request.POST.get('balance')
        premium = request.POST.get('premium')
        other = request.POST.get('other')
        car = request.POST.get('car')
        journey = request.POST.get('journey')
        payement_method = request.POST.get('payement_method')
        period_pattern = r'^\d{2}/\d{2}/\d{4} au \d{2}/\d{2}/\d{4}$'
        if not re.match(period_pattern, period):
            context['error_message'] = "La période doit être sous forme de 'jj/mm/aaaa au jj/mm/aaaa'."
            return render(request, 'insurance/add_insurance.html', context)
        
        if insurance_type == "Santé":
            insurance = Insurance.objects.create(
                date = datetime.now(),
                insurance_type=insurance_type,
                certificate_number=certificate_number,
                policy_number=policy_number,
                insured=insured,
                period=period,
                address=address,
                phone=phone,
                deposit=deposit,
                balance=balance,
                premium=premium,
                payement_method=payement_method,
                designation=""
            )
        elif insurance_type == "Voyage":
            insurance = Insurance.objects.create(
                date = datetime.now(),
                insurance_type=insurance_type,
                certificate_number=certificate_number,
                policy_number=policy_number,
                insured=insured,
                period=period,
                address=address,
                phone=phone,
                deposit=deposit,
                balance=balance,
                premium=premium,
                payement_method=payement_method,
                designation=journey
            )
        elif insurance_type == "Auto":
            insurance = Insurance.objects.create(
                date = datetime.now(),
                insurance_type=insurance_type,
                certificate_number=certificate_number,
                policy_number=policy_number,
                insured=insured,
                period=period,
                address=address,
                phone=phone,
                deposit=deposit,
                balance=balance,
                premium=premium,
                payement_method=payement_method,
                designation=car
            )
        else:
            insurance = Insurance.objects.create(
                date = datetime.now(),
                insurance_type=insurance_type,
                certificate_number=certificate_number,
                policy_number=policy_number,
                insured=insured,
                period=period,
                address=address,
                phone=phone,
                deposit=deposit,
                balance=balance,
                premium=premium,
                payement_method=payement_method,
                designation=other
            )
        insurance.save()
        return redirect('view_insurance')
    return render(request, 'insurance/add_insurance.html', context)

@permission_required('amana.change_insurance', raise_exception=True)
def payment_insurance(request, insurance_id):
    insurance = get_object_or_404(Insurance, id=insurance_id)
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'insurance': insurance,
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
    }
    if request.method == 'POST':
        payment = request.POST.get('payment')
        try:
            payment_decimal = Decimal(payment)
            insurance.deposit += payment_decimal
            insurance.balance = insurance.premium - insurance.deposit
            insurance.date = datetime.now()
            insurance.save()
            return redirect('view_insurance')
        except (Decimal.InvalidOperation, ValueError):
            context['error'] = "Invalid payment amount."
            return render(request, 'insurance/payment_insurance.html', context)
        
    return render(request, 'insurance/payment_insurance.html', context)

@permission_required('amana.change_insurance', raise_exception=True)
def change_insurance(request, insurance_id):
    insurance = get_object_or_404(Insurance, id=insurance_id)
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'insurance': insurance,
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
    }
    if request.method == 'POST':
        certificate_number = request.POST.get('certificate_number')
        policy_number = request.POST.get('policy_number')
        insured = request.POST.get('insured')
        period = request.POST.get('period')
        address = request.POST.get('address')
        phone = request.POST.get('phone')
        insurance_type = request.POST.get('insurance_type')
        phone = request.POST.get('phone')
        deposit = request.POST.get('deposit')
        balance = request.POST.get('balance')
        premium = request.POST.get('premium')
        other = request.POST.get('other')
        car = request.POST.get('car')
        journey = request.POST.get('journey')
        payement_method = request.POST.get('payement_method')
        period_pattern = r'^\d{2}/\d{2}/\d{4} au \d{2}/\d{2}/\d{4}$'
        if not re.match(period_pattern, period):
            context['error_message'] = "La période doit être sous forme de 'jj/mm/aaaa au jj/mm/aaaa'."
            return render(request, 'insurance/change_insurance.html', context)
        if insurance_type == "Voyage":
            insurance.designation=journey
        elif insurance_type == "Auto":
            insurance.designation=car
        elif insurance_type == "Autre":
            insurance.designation=other
        insurance.certificate_number=certificate_number
        insurance.policy_number=policy_number
        insurance.insured=insured
        insurance.period=period
        insurance.address=address
        insurance.phone=phone
        insurance.insurance_type=insurance_type
        insurance.deposit=deposit
        insurance.balance=balance
        insurance.premium=premium
        insurance.date = datetime.now()
        insurance.payement_method=payement_method
        insurance.save()
        return redirect('view_insurance')
    return render(request, 'insurance/change_insurance.html', context)

@permission_required('amana.delete_insurance', raise_exception=True)
def delete_insurance(request, insurance_id):
    insurance = get_object_or_404(Insurance, id=insurance_id)
    insurance.delete()
    return redirect('view_insurance')

@permission_required('amana.view_insurance', raise_exception=True)
def insurance(request):
    payment_method = request.GET.get('payment_method', '')
    search_date = request.GET.get('date', '')
    insurances = Insurance.objects.all()
    insurances_dates = Insurance.objects.annotate(
        date_only=Cast('date', output_field=DateField())
    ).values_list('date_only', flat=True).distinct().order_by('date_only')
    if search_date:
        locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')
        search_date = datetime.strptime(search_date, '%d %B %Y').date()
        insurances = insurances.annotate(
            date_only=TruncDate('date')
        ).filter(date_only=search_date)
    if payment_method:
        insurances = insurances.filter(payement_method=payment_method)
    
    totals = insurances.aggregate(
        total_balance=Sum(
            Case(
                When(payement_method="En euro", then=F('balance') * 490),
                default=F('balance'),
                output_field=IntegerField(),
            )
        ),
        total_deposit=Sum(
            Case(
                When(payement_method="En euro", then=F('deposit') * 490),
                default=F('deposit'),
                output_field=IntegerField(),
            )
        ),
        total_premium=Sum(
            Case(
                When(payement_method="En euro", then=F('premium') * 490),
                default=F('premium'),
                output_field=IntegerField(),
            )
        )
    )

    total_balance = totals.get('total_balance') or 0
    total_deposit = totals.get('total_deposit') or 0
    total_premium = totals.get('total_premium') or 0

    insurances = insurances.order_by('-id')

    context = {
        'insurances': insurances,
        'total_balance': total_balance,
        'total_deposit': total_deposit,
        'total_premium': total_premium,
        'payment_method': payment_method,
        'insurances_dates': insurances_dates,
        'search_date':search_date
    }

    return render(request, 'insurance/view.html', context)

@permission_required('amana.view_smallcrate', raise_exception=True)
def view_smallcrate(request):
    search_query = request.POST.get('search', '')
    smallcrate = SmallCrate.objects.all()
    if search_query:
        smallcrate = smallcrate.filter(designation__icontains=search_query)
    smallcrate = smallcrate.order_by('-id')
    paginator = Paginator(smallcrate, 12) 
    page_number = request.GET.get('page', 1) 
    page_smallcrate = paginator.get_page(page_number)
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
       'smallcrates': page_smallcrate,
       'search_query':search_query
    }
    return render(request, 'disbursement/smallcrate/view_smallcrate.html', context)

@permission_required('amana.add_smallcrate', raise_exception=True)
def add_smallcrate(request):
    if request.method == 'POST':
        designation = request.POST.get('designation')
        amount = request.POST.get('amount')
        disbursement = Disbursement.objects.create(
            date = datetime.now(),
            amount=amount
        )
        disbursement.save()
        smallcrate = SmallCrate.objects.create(
            designation = designation,
            disbursement=disbursement,
        )
        smallcrate.save()
        return redirect('view_smallcrate')
    
    current_user = request.user
    first_name_current = current_user.first_name.capitalize()
    last_name_current = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name_current[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'first_name': first_name_current,
        'last_name': last_name_current,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
    }
    return render(request, 'disbursement/smallcrate/add_smallcrate.html', context)

@permission_required('amana.delete_smallcrate', raise_exception=True)
def delete_smallcrate(request, smallcrate_id):
    smallcrate = get_object_or_404(SmallCrate, id=smallcrate_id)
    disbursement = get_object_or_404(Disbursement, id=smallcrate.disbursement.id)
    smallcrate.delete()
    disbursement.delete()
    return redirect('view_smallcrate')

@permission_required('amana.change_smallcrate', raise_exception=True)
def change_smallcrate(request, smallcrate_id):
    smallcrate = get_object_or_404(SmallCrate, id=smallcrate_id)
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'smallcrate': smallcrate,
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
    }
    if request.method == 'POST':
        designation = request.POST.get('designation')
        amount = request.POST.get('amount')
        smallcrate.designation = designation
        smallcrate.save()
        disbursement = get_object_or_404(Disbursement, id=smallcrate.disbursement.id)
        disbursement.date = datetime.now()
        disbursement.amount = amount
        disbursement.save()
        return redirect('view_smallcrate')
    return render(request, 'disbursement/smallcrate/change_smallcrate.html', context)

@permission_required('amana.view_health', raise_exception=True)
def view_health(request):
    search_query = request.POST.get('search', '')
    healths = Health.objects.all()
    if search_query:
        healths = healths.filter(registration_number__icontains=search_query)
    healths = healths.order_by('-id')
    paginator_healths = Paginator(healths, 12) 
    page_number = request.GET.get('page', 1) 
    page_healths = paginator_healths.get_page(page_number)
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
       'healths': page_healths,
       'search_query':search_query
    }
    return render(request, 'disbursement/health/view_health.html', context)

@permission_required('amana.add_health', raise_exception=True)
def add_health(request):
    if request.method == 'POST':
        registration_number = request.POST.get('registration_number')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        company = request.POST.get('company')
        status = request.POST.get('status')
        amount = request.POST.get('amount')
        disbursement = Disbursement.objects.create(
            date = datetime.now(),
            amount=amount
        )
        disbursement.save()
        health = Health.objects.create(
            registration_number = registration_number,
            first_name=first_name,
            last_name=last_name,
            company=company,
            status=status,
            disbursement=disbursement,
        )
        health.save()
        return redirect('view_health')
    
    current_user = request.user
    first_name_current = current_user.first_name.capitalize()
    last_name_current = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name_current[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'first_name': first_name_current,
        'last_name': last_name_current,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
    }
    return render(request, 'disbursement/health/add_health.html', context)

@permission_required('amana.delete_health', raise_exception=True)
def delete_health(request, health_id):
    health = get_object_or_404(Health, id=health_id)
    disbursement = get_object_or_404(Disbursement, id=health.disbursement.id)
    health.delete()
    disbursement.delete()
    return redirect('view_health')

@permission_required('amana.change_health', raise_exception=True)
def change_health(request, health_id):
    health = get_object_or_404(Health, id=health_id)
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'health': health,
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
    }
    if request.method == 'POST':
        registration_number = request.POST.get('registration_number')
        amana_first_name = request.POST.get('first_name')
        amana_last_name = request.POST.get('last_name')
        company = request.POST.get('company')
        status = request.POST.get('status')
        amount = request.POST.get('amount')
        health.first_name = amana_first_name
        health.last_name = amana_last_name
        health.company = company
        health.status = status
        health.registration_number = registration_number
        health.save()
        disbursement = get_object_or_404(Disbursement, id=health.disbursement.id)
        disbursement.date = datetime.now()
        disbursement.amount = amount
        disbursement.save()
        return redirect('view_health')
    return render(request, 'disbursement/health/change_health.html', context)

@permission_required('amana.view_journey', raise_exception=True)
def view_journey(request):
    search_query = request.POST.get('search', '')
    journey = Journey.objects.all()
    if search_query:
        journey = journey.filter(registration_number__icontains=search_query)
    journey = journey.order_by('-id')
    paginator_journey = Paginator(journey, 12) 
    page_number = request.GET.get('page', 1) 
    page_journey = paginator_journey.get_page(page_number)
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
       'journeys':page_journey,
       'search_query':search_query
    }
    return render(request, 'disbursement/journey/view_journey.html', context)

@permission_required('amana.add_journey', raise_exception=True)
def add_journey(request):
    if request.method == 'POST':
        registration_number = request.POST.get('registration_number')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        amount = request.POST.get('amount')
        disbursement = Disbursement.objects.create(
            date = datetime.now(),
            amount=amount
        )
        disbursement.save()
        amana = Journey.objects.create(
            registration_number = registration_number,
            first_name=first_name,
            last_name=last_name,
            disbursement=disbursement,
        )
        amana.save()
        return redirect('view_journey')
    
    current_user = request.user
    first_name_current = current_user.first_name.capitalize()
    last_name_current = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name_current[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'first_name': first_name_current,
        'last_name': last_name_current,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
    }
    return render(request, 'disbursement/journey/add_journey.html', context)

@permission_required('amana.delete_journey', raise_exception=True)
def delete_journey(request, journey_id):
    journey = get_object_or_404(Journey, id=journey_id)
    disbursement = get_object_or_404(Disbursement, id=journey.disbursement.id)
    journey.delete()
    disbursement.delete()
    return redirect('view_journey')

@permission_required('amana.change_journey', raise_exception=True)
def change_journey(request, journey_id):
    journey = get_object_or_404(Journey, id=journey_id)
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'journey': journey,
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
    }
    if request.method == 'POST':
        registration_number = request.POST.get('registration_number')
        journey_first_name = request.POST.get('first_name')
        journey_last_name = request.POST.get('last_name')
        amount = request.POST.get('amount')
        journey.first_name = journey_first_name
        journey.last_name = journey_last_name
        journey.registration_number = registration_number
        journey.save()
        disbursement = get_object_or_404(Disbursement, id=journey.disbursement.id)
        disbursement.amount = amount
        disbursement.date = datetime.now()
        disbursement.save()
        return redirect('view_journey')
    return render(request, 'disbursement/journey/change_journey.html', context)

@permission_required('amana.view_car', raise_exception=True)
def view_car(request):
    search_query = request.POST.get('search', '')
    car = Car.objects.all()
    if search_query:
        car = car.filter(Q(claim_number__icontains=search_query) | Q(policy_number__icontains=search_query))
    car = car.order_by('-id')
    paginator_car = Paginator(car, 12) 
    page_number = request.GET.get('page', 1) 
    page_car = paginator_car.get_page(page_number)
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
       'cars':page_car,
       'search_query':search_query
    }
    return render(request, 'disbursement/car/view_car.html', context)

@permission_required('amana.add_car', raise_exception=True)
def add_car(request):
    if request.method == 'POST':
        claim_number = request.POST.get('claim_number')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        designation = request.POST.get('designation')
        payement_method = request.POST.get('payement_method')
        policy_number = request.POST.get('policy_number')
        amount = request.POST.get('amount')
        disbursement = Disbursement.objects.create(
            date = datetime.now(),
            amount=amount
        )
        disbursement.save()
        amana = Car.objects.create(
            claim_number = claim_number,
            first_name=first_name,
            last_name=last_name,
            designation=designation,
            payement_method=payement_method,
            policy_number=policy_number,
            disbursement=disbursement,
        )
        amana.save()
        return redirect('view_car')
    
    current_user = request.user
    first_name_current = current_user.first_name.capitalize()
    last_name_current = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name_current[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'first_name': first_name_current,
        'last_name': last_name_current,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
    }
    return render(request, 'disbursement/car/add_car.html', context)

@permission_required('amana.delete_car', raise_exception=True)
def delete_car(request, car_id):
    car = get_object_or_404(Car, id=car_id)
    disbursement = get_object_or_404(Disbursement, id=car.disbursement.id)
    car.delete()
    disbursement.delete()
    return redirect('view_car')

@permission_required('amana.change_car', raise_exception=True)
def change_car(request, car_id):
    car = get_object_or_404(Car, id=car_id)
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'car': car,
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
    }
    if request.method == 'POST':
        claim_number = request.POST.get('claim_number')
        car_first_name = request.POST.get('first_name')
        car_last_name = request.POST.get('last_name')
        designation = request.POST.get('designation')
        payement_method = request.POST.get('payement_method')
        policy_number = request.POST.get('policy_number')
        amount = request.POST.get('amount')
        car.first_name = car_first_name
        car.last_name = car_last_name
        car.claim_number = claim_number
        car.payement_method = payement_method
        car.designation = designation
        car.policy_number = policy_number
        car.save()
        disbursement = get_object_or_404(Disbursement, id=car.disbursement.id)
        disbursement.amount = amount
        disbursement.date = datetime.now()
        disbursement.save()
        return redirect('view_car')
    return render(request, 'disbursement/car/change_car.html', context)

@permission_required('amana.view_production', raise_exception=True)
def view_production(request):
    context = {}
    if request.method == 'POST':
        production_type = request.POST.get('production')
        selected_month = int(request.POST.get('month'))
        selected_year = int(request.POST.get('year'))

        insurances = Insurance.objects.filter(insurance_type=production_type)
        filtered_insurances = []

        for insurance in insurances:
            try:
                start_date_str = insurance.period.split()[0]
                start_date = datetime.strptime(start_date_str, '%d/%m/%Y')
                insurance_month = start_date.month
                insurance_year = start_date.year

                if insurance_month == selected_month and insurance_year == selected_year and int(insurance.balance) == 0:
                    if insurance.payement_method == "En euro":
                        insurance.deposit = int(insurance.deposit) * 490
                    filtered_insurances.append(insurance)
            except ValueError:
                continue
        
        if filtered_insurances:
            grey_fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
            workbook = Workbook()
            sheet = workbook.active

            if production_type == "Auto":
                sheet.title = f"Production {production_type}"
                line_title = [
                    "", ""," ", f"PRODUCTION {production_type.upper()}", "", "", "", ""
                ]
                headers = [
                    "POLICE", "PERIODE"," ", "ASSURES", "PRIMES", "ACCES-", "TAXE 4%", "PRIMES"
                ]
            elif production_type == "Autre":
                sheet.title = f"Production {production_type}"
                line_title = [
                    "", ""," ", f"PRODUCTION {production_type.upper()}", "", "", ""
                ]
                headers = [
                    "POLICE", "PERIODE"," ", "ASSURES", "PRIMES", "TAXE 4%", "PRIMES"
                ]
            elif production_type == "Santé":
                sheet.title = "PRODUCTION SANTE"
                headers = [
                    "POLICE", "PERIODE"," ", "ASSURES", "PRIMES", "TAXE 3%", "PRIMES"
                ]
                line_title = [
                    "", ""," ", f"PRODUCTION {production_type.upper()}", "", "", ""
                ]
            else:
                sheet.title = "Production Voyage"
                headers = [
                    "POLICE", "PERIODE"," ", "ASSURES", "PRIMES"
                ]
                line_title = [
                    "", ""," ", "PRODUCTION VOYAGE", ""
                ]

            sheet.append([])
            sheet.append(line_title)
            sheet.append([])
            sheet.append(headers)

            if production_type == "Auto":
                sheet.merge_cells('A4:A5')
                sheet.merge_cells('B4:C4')
                sheet.merge_cells('D4:D5')
                sheet['B5'] = 'du'
                sheet['C5'] = 'au'

                sheet.merge_cells('G4:G5')
                sheet['F5'] = 'SOIRES'
                sheet['E5'] = 'NETTES'
                sheet['H5'] = 'TOTALES'
            elif production_type == "Santé" or production_type == "Autre":
                sheet.merge_cells('A4:A5')
                sheet.merge_cells('B4:C4')
                sheet.merge_cells('D4:D5')
                sheet['B5'] = 'du'
                sheet['C5'] = 'au'
                
                sheet.merge_cells('F4:F5')
                sheet['E5'] = 'NETTES'
                sheet['G5'] = 'TOTALES'
            else:
                sheet.merge_cells('A4:A5')
                sheet.merge_cells('B4:C4')
                sheet.merge_cells('D4:D5')
                sheet['B5'] = 'du'
                sheet['C5'] = 'au'
                sheet['E5'] = 'NETTES'

            for row in sheet.iter_rows(min_row=2, max_row=2):
                for cell in row:
                    cell.font = Font(bold=True, size=16)  
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            for row in sheet.iter_rows(min_row=4, max_row=5):
                for cell in row:
                    cell.fill = grey_fill
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            column_widths = {
                "A": 128 / 7, 
                "B": 128 / 7,
                "C": 128 / 7,
                "D": 702 / 7,
                "E": 128 / 7,
                "F": 128 / 7,
                "G": 128 / 7,
                "H": 128 / 7,
            }

            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

            sheet.auto_filter.ref = f"A5:H5" if production_type == "Auto" else (
                f"A5:G5" if production_type in ["Santé", "Autre"] else f"A5:E5"
            )

            total_taxe, total_accessories, total_deposit, total_total = 0, 0, 0, 0

            for insurance in filtered_insurances:
                start_date_str, end_date_str = insurance.period.split()[0], insurance.period.split()[2]
                accessories, taxe, total = 0, 0, 0

                if production_type == "Auto":
                    accessories = 2000
                    taxe = int(int(insurance.deposit) * 0.04)
                    total = int(insurance.deposit) + taxe + accessories
                    total_taxe += taxe
                    total_accessories += accessories
                elif production_type == "Santé":
                    taxe = int(int(insurance.deposit) * 0.03)
                    total = int(insurance.deposit) + taxe
                    total_taxe += taxe
                elif production_type == "Autre":
                    taxe = int(int(insurance.deposit) * 0.04)
                    total = int(insurance.deposit) + taxe
                    total_taxe += taxe

                total_deposit += int(insurance.deposit)
                total_total += total

                if production_type == "Auto":
                    row = [
                        insurance.policy_number, 
                        f"{start_date_str}",
                        f"{end_date_str}",
                        insurance.insured,
                        insurance.deposit,
                        accessories,
                        taxe,
                        total,
                    ]
                elif production_type == "Santé" or production_type == "Autre":
                    row = [
                        insurance.policy_number, 
                        f"{start_date_str}",
                        f"{end_date_str}",
                        insurance.insured,
                        insurance.deposit,
                        taxe,
                        total,
                    ]
                else:
                    row = [
                        insurance.policy_number, 
                        f"{start_date_str}",
                        f"{end_date_str}",
                        insurance.insured,
                        insurance.deposit,
                    ]
                sheet.append(row)
            
            for row in sheet.iter_rows(min_row=4, min_col=2, max_col=3):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            if production_type == "Auto":
                line_total = [
                    "", ""," ", "TOTAUX......................................", f"{total_deposit}", f"{total_accessories}", f"{total_taxe}", f"{total_total}"
                ]
            
            elif production_type == "Santé" or production_type == "Autre":
                line_total = [
                    "", ""," ", "TOTAUX......................................", f"{total_deposit}" , f"{total_taxe}", f"{total_total}"
                ]
            else:
                line_total = [
                    "", ""," ", "TOTAUX......................................", f"{total_deposit}"
                ]
            sheet.append([])
            sheet.append([])
            sheet.append([])
            sheet.append(line_total)

            for row in sheet.iter_rows(min_row=sheet.max_row, min_col=4):
                for cell in row:
                    cell.fill = grey_fill
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='right', vertical='center')

            excel_data = save_virtual_workbook(workbook)

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="production_{production_type}_{selected_month}_{selected_year}.xlsx"'
            response.write(excel_data)

            return response
        else:
            context['error'] = 'Aucune assurance n\'a été payée dans la période choisie.'

    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context.update({
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
    })

    return render(request, 'production/production.html', context)

@permission_required('amana.view_primeassurance', raise_exception=True)
def view_primeassurance(request):
    search_query = request.POST.get('search', '')
    primeassurance = PrimeAssurance.objects.all()
    
    if search_query:
        primeassurance = primeassurance.filter(responsible__icontains=search_query)
    
    primeassurance_data = []
    for assurance in primeassurance:
        num_insured = Insured.objects.filter(responsible=assurance.responsible).values('registration_number').distinct().count()
        primeassurance_data.append({
            'id': assurance.id,
            'date_emission': assurance.date_emission,
            'date_fin': assurance.date_fin,
            'prime': assurance.prime,
            'responsible': assurance.responsible,
            'num_insured': num_insured
        })
    
    primeassurance_data.sort(key=lambda x: x['id'], reverse=True)
    paginator = Paginator(primeassurance_data, 12)
    page_number = request.GET.get('page', 1)
    page = paginator.get_page(page_number)
    
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    
    context = {
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
        'primeassurances': page,
        'search_query': search_query
    }
    
    return render(request, 'primeassurance/view_primeassurance.html', context)

@permission_required('amana.add_primeassurance', raise_exception=True)
def add_primeassurance(request):
    if request.method == 'POST':
        date_emission = request.POST.get('date_emission')
        date_fin = request.POST.get('date_fin')
        prime = request.POST.get('prime')
        responsible = request.POST.get('responsible')
        primeassurance = PrimeAssurance.objects.create(
            date_emission = date_emission,
            date_fin=date_fin,
            prime=prime,
            responsible=responsible
        )
        primeassurance.save()
        return redirect('view_primeassurance')
    
    current_user = request.user
    first_name_current = current_user.first_name.capitalize()
    last_name_current = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name_current[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'first_name': first_name_current,
        'last_name': last_name_current,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
    }
    return render(request, 'primeassurance/add_primeassurance.html', context)

@permission_required('amana.change_primeassurance', raise_exception=True)
def change_primeassurance(request, primeassurance_id):
    primeassurance = get_object_or_404(PrimeAssurance, id=primeassurance_id)
    old_responsible= primeassurance.responsible
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'primeassurance': primeassurance,
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
    }
    
    if request.method == 'POST':
        date_emission = request.POST.get('date_emission')
        date_fin = request.POST.get('date_fin')
        prime = request.POST.get('prime')
        responsible = request.POST.get('responsible')
        
        primeassurance.date_emission = date_emission
        primeassurance.date_fin = date_fin
        primeassurance.prime = prime
        primeassurance.responsible = responsible
        primeassurance.save()
        
        if old_responsible == responsible:
            insured_count = Insured.objects.filter(responsible=responsible).count()
            if insured_count > 0:
                new_prime_per_insured = int(prime) / insured_count
                Insured.objects.filter(responsible=responsible).update(prime=new_prime_per_insured)
        
        else:  
            Insured.objects.filter(responsible=old_responsible).update(responsible=responsible)
            
            insured_count_new = Insured.objects.filter(responsible=responsible).count()
            if insured_count_new > 0:
                new_prime_per_insured_new = int(prime) / insured_count_new
                Insured.objects.filter(responsible=responsible).update(prime=new_prime_per_insured_new)
        
        return redirect('view_primeassurance')
    
    return render(request, 'primeassurance/change_primeassurance.html', context)

@permission_required('amana.view_insured', raise_exception=True)
def primeassurance_view_insured(request, primeassurance_id):
    search_query = request.POST.get('search', '')
    primeassurance = get_object_or_404(PrimeAssurance, id=primeassurance_id)
    insured = Insured.objects.filter(responsible=primeassurance.responsible)
    if search_query:
        insured = insured.filter(Q(insured__icontains=search_query) | Q(registration_number__icontains=search_query))
    insured = insured.order_by('-registration_number')
    paginator = Paginator(insured, 12)
    page_number = request.GET.get('page', 1)
    page = paginator.get_page(page_number)
    
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    
    context = {
        'primeassurance_id':primeassurance_id,
        'responsible':primeassurance.responsible,
        'insureds':page,
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
        'insured': page,
        'search_query': search_query
    }
    
    return render(request, 'primeassurance/view_insured.html', context)

@permission_required('amana.add_insured', raise_exception=True)
def primeassurance_add_insured(request, primeassurance_id):
    primeassurance = get_object_or_404(PrimeAssurance, id=primeassurance_id)
    responsible = primeassurance.responsible
    existing_registration_numbers = Insured.objects.filter(responsible=responsible).values_list('registration_number', flat=True).distinct()

    if request.method == 'POST':
        registration_number = request.POST.get('registration_number')
        name_insured = request.POST.get('insured')
        status = request.POST.get('status')

        with transaction.atomic():
            insured_ct = len(existing_registration_numbers)
            if registration_number in existing_registration_numbers:
                insured = Insured.objects.create(
                    registration_number=registration_number,
                    insured=name_insured,
                    prime=0,
                    status=status,
                    responsible=responsible
                )
            else:
                # Check if there are existing insured entries
                if insured_ct > 0:
                    prime_per_existing_insured = int(primeassurance.prime) / insured_ct
                    primeassurance.prime = int(primeassurance.prime) + prime_per_existing_insured
                else:
                    prime_per_existing_insured = 0
                
                # Create new Insured entry
                insured = Insured.objects.create(
                    registration_number=registration_number,
                    insured=name_insured,
                    prime=prime_per_existing_insured,
                    status=status,
                    responsible=responsible
                )
                insured.save()

            # Recalculate the prime per insured
            insured_count = Insured.objects.filter(responsible=responsible).count()
            if insured_count > 0:
                prime_per_insured = int(primeassurance.prime) / insured_count
            else:
                prime_per_insured = 0

            Insured.objects.filter(responsible=responsible).update(prime=prime_per_insured)
            primeassurance.save()

        return redirect('primeassurance_view_insured', primeassurance_id=primeassurance_id)

    current_user = request.user
    first_name_current = current_user.first_name.capitalize()
    last_name_current = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name_current[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'

    context = {
        'responsible': responsible,
        'first_name': first_name_current,
        'last_name': last_name_current,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
    }

    return render(request, 'primeassurance/add_insured.html', context)

@permission_required('amana.change_insured', raise_exception=True)
def primeassurance_change_insured(request, primeassurance_id,insured_id):
    insured = get_object_or_404(Insured, id=insured_id)
    old_status = insured.status
    old_responsible = insured.responsible
    old_registration_number = insured.registration_number
    old_primeassurance = get_object_or_404(PrimeAssurance, id=primeassurance_id)
    if request.method == 'POST':
        registration_number = request.POST.get('registration_number')
        name_insured = request.POST.get('insured')
        status = request.POST.get('status')
        
        with transaction.atomic():
            if old_status == "Assuré":
                new_insured_count = Insured.objects.filter(responsible=old_responsible, status="Assuré").count()
                insured.registration_number = registration_number
                insured.insured = name_insured
                insured.status = status
                insured.save()
                if status != "Assuré":
                    if new_insured_count > 0:
                        prime = int(old_primeassurance.prime) / int(new_insured_count)
                        old_primeassurance.prime = int(old_primeassurance.prime) - int(prime)
                        old_primeassurance.save()
                        Insured.objects.filter(responsible=old_responsible, registration_number=old_registration_number).delete()
                        insured_count = Insured.objects.filter(responsible=old_responsible).count()
                        prime_per_insured = int(old_primeassurance.prime) / int(insured_count)
                        Insured.objects.filter(responsible=old_responsible).update(prime=prime_per_insured)
            else:
                new_insured_count = Insured.objects.filter(responsible=old_responsible, status="Assuré").count()
                insured.registration_number = registration_number
                insured.insured = name_insured
                insured.status = status
                insured.save()
                if status == "Assuré":
                    if new_insured_count > 0:
                        prime = int(old_primeassurance.prime) / int(new_insured_count)
                        old_primeassurance.prime = int(old_primeassurance.prime) + int(prime)
                        old_primeassurance.save()
                        insured_count = Insured.objects.filter(responsible=old_responsible).count()
                        prime_per_insured = int(old_primeassurance.prime) / int(insured_count)
                        Insured.objects.filter(responsible=old_responsible).update(prime=prime_per_insured)
        return redirect('primeassurance_view_insured', primeassurance_id=primeassurance_id)      
    current_user = request.user
    first_name_current = current_user.first_name.capitalize()
    last_name_current = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name_current[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    
    context = {
        'first_name': first_name_current,
        'last_name': last_name_current,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
        'insured': insured
    }
    
    return render(request, 'primeassurance/change_insured.html', context)

@permission_required('amana.delete_insured', raise_exception=True)
def primeassurance_delete_insured(request, primeassurance_id, insured_id):
    primeassurance = get_object_or_404(PrimeAssurance, id=primeassurance_id)
    insured = get_object_or_404(Insured, id=insured_id, responsible=primeassurance.responsible)

    if insured.status == "Assuré":
        registration_number = insured.registration_number
        insured_total_count = Insured.objects.filter(responsible=primeassurance.responsible).values('registration_number').distinct().count()
        old_insured_count = Insured.objects.filter(responsible=primeassurance.responsible, registration_number=insured.registration_number).values('registration_number').distinct().count()
        insured.delete()
        existing_insured = Insured.objects.filter(responsible=primeassurance.responsible, registration_number=registration_number)
        if existing_insured:
            existing_insured.delete()
        if old_insured_count > 0 and insured_total_count > 0:
            count = insured_total_count - old_insured_count
            prime_per_insured = (int(primeassurance.prime) * count) / insured_total_count
            primeassurance.prime = prime_per_insured
            primeassurance.save()
        insured_count = Insured.objects.filter(responsible=primeassurance.responsible).count()
        if insured_count > 0:
            prime_per_insured = int(primeassurance.prime) / insured_count
            Insured.objects.filter(responsible=primeassurance.responsible).update(prime=prime_per_insured)
    
    else:
        insured.delete()
        insured_count = Insured.objects.filter(responsible=primeassurance.responsible).count()
        if insured_count > 0:
            prime_per_insured = int(primeassurance.prime) / insured_count
            Insured.objects.filter(responsible=primeassurance.responsible).update(prime=prime_per_insured)

    return redirect('primeassurance_view_insured', primeassurance_id=primeassurance_id)

@permission_required('amana.delete_primeassurance', raise_exception=True)
def delete_primeassurance(request, primeassurance_id):
    primeassurance = get_object_or_404(PrimeAssurance, id=primeassurance_id)
    insured = Insured.objects.filter(responsible=primeassurance.responsible)
    insured.delete()
    primeassurance.delete()
    return redirect('view_primeassurance')

@permission_required('amana.view_insured', raise_exception=True)
def view_insured(request):
    search_query = request.POST.get('search', '')
    insured = Insured.objects.all()
    if search_query:
        insured = insured.filter(Q(responsible__icontains=search_query) | Q(insured__icontains=search_query) | Q(registration_number__icontains=search_query))
    insured = insured.order_by('-registration_number')
    paginator = Paginator(insured, 12) 
    page_number = request.GET.get('page', 1) 
    page = paginator.get_page(page_number)
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
       'insureds':page,
       'search_query':search_query
    }
    return render(request, 'insured/view_insured.html', context)


@permission_required('amana.add_insured', raise_exception=True)
def add_insured(request):
    if request.method == 'POST':
        registration_number = request.POST.get('registration_number')
        name_insured = request.POST.get('insured')
        responsible = request.POST.get('responsible')
        status = request.POST.get('status')
        
        existing_registration_numbers = Insured.objects.filter(responsible=responsible).values_list('registration_number', flat=True).distinct()
        old_primeassurance = get_object_or_404(PrimeAssurance, responsible=responsible)

        with transaction.atomic():
            insured_ct = len(existing_registration_numbers)
            if registration_number in existing_registration_numbers:
                insured = Insured.objects.create(
                    registration_number=registration_number,
                    insured=name_insured,
                    prime=0,
                    status=status,
                    responsible=responsible
                )
            else:
                if insured_ct > 0:
                    prime_per_existing_insured = int(old_primeassurance.prime) / insured_ct
                    old_primeassurance.prime = int(old_primeassurance.prime) + prime_per_existing_insured
                else:
                    prime_per_existing_insured = 0
                
                insured = Insured.objects.create(
                    registration_number=registration_number,
                    insured=name_insured,
                    prime=prime_per_existing_insured,
                    status=status,
                    responsible=responsible
                )
                insured.save()
            
            insured_count = Insured.objects.filter(responsible=responsible).count()
            if insured_count > 0:
                prime_per_insured = int(old_primeassurance.prime) / insured_count
            else:
                prime_per_insured = 0

            Insured.objects.filter(responsible=responsible).update(prime=prime_per_insured)
            old_primeassurance.save()
        
        return redirect('view_insured')
    
    primeassurance = PrimeAssurance.objects.all()
    current_user = request.user
    first_name_current = current_user.first_name.capitalize()
    last_name_current = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name_current[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    
    context = {
        'first_name': first_name_current,
        'last_name': last_name_current,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
        'primeassurance': primeassurance,
    }
    
    return render(request, 'insured/add_insured.html', context)

@permission_required('amana.change_insured', raise_exception=True)
def change_insured(request, insured_id):
    insured = get_object_or_404(Insured, id=insured_id)
    old_primeassurance = PrimeAssurance.objects.filter(responsible=insured.responsible).first()
    primeassurance = PrimeAssurance.objects.all()
    old_responsible = insured.responsible
    old_registration_number = insured.registration_number
    old_status = insured.status 
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'

    context = {
        'primeassurance': primeassurance,
        'insured': insured,
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
    }
    
    if request.method == 'POST':
        registration_number = request.POST.get('registration_number')
        name_insured = request.POST.get('insured')
        responsible = request.POST.get('responsible')
        status = request.POST.get('status')

        with transaction.atomic():
            if old_responsible == responsible:
                if old_status == "Assuré":
                    new_insured_count = Insured.objects.filter(responsible=old_responsible, status="Assuré").count()
                    insured.registration_number = registration_number
                    insured.insured = name_insured
                    insured.responsible = responsible
                    insured.status = status
                    insured.save()
                    if status != "Assuré":
                        if new_insured_count > 0:
                            prime = int(old_primeassurance.prime) / int(new_insured_count)
                            old_primeassurance.prime = int(old_primeassurance.prime) - int(prime)
                            old_primeassurance.save()
                            Insured.objects.filter(responsible=old_responsible, registration_number=old_registration_number).delete()
                            insured_count = Insured.objects.filter(responsible=old_responsible).count()
                            prime_per_insured = int(old_primeassurance.prime) / int(insured_count)
                            Insured.objects.filter(responsible=old_responsible).update(prime=prime_per_insured)
                else:
                    new_insured_count = Insured.objects.filter(responsible=old_responsible, status="Assuré").count()
                    insured.registration_number = registration_number
                    insured.insured = name_insured
                    insured.responsible = responsible
                    insured.status = status
                    insured.save()
                    if status == "Assuré":
                        if new_insured_count > 0:
                            prime = int(old_primeassurance.prime) / int(new_insured_count)
                            old_primeassurance.prime = int(old_primeassurance.prime) + int(prime)
                            old_primeassurance.save()
                            insured_count = Insured.objects.filter(responsible=old_responsible).count()
                            prime_per_insured = int(old_primeassurance.prime) / int(insured_count)
                            Insured.objects.filter(responsible=old_responsible).update(prime=prime_per_insured)
                        
        return redirect('view_insured')

    return render(request, 'insured/change_insured.html', context)

@permission_required('amana.delete_insured', raise_exception=True)
def delete_insured(request, insured_id):
    insured = get_object_or_404(Insured, id=insured_id)
    primeassurance = get_object_or_404(PrimeAssurance, responsible=insured.responsible)
    if insured.status == "Assuré":
        registration_number = insured.registration_number
        insured_total_count = Insured.objects.filter(responsible=primeassurance.responsible).values('registration_number').distinct().count()
        old_insured_count = Insured.objects.filter(responsible=primeassurance.responsible, registration_number=insured.registration_number).values('registration_number').distinct().count()
        insured.delete()
        existing_insured = Insured.objects.filter(responsible=primeassurance.responsible, registration_number=registration_number)
        if existing_insured:
            existing_insured.delete()
        if old_insured_count > 0 and insured_total_count > 0:
            count = insured_total_count - old_insured_count
            prime_per_insured = (int(primeassurance.prime) * count) / insured_total_count
            primeassurance.prime = prime_per_insured
            primeassurance.save()
        insured_count = Insured.objects.filter(responsible=primeassurance.responsible).count()
        if insured_count > 0:
            prime_per_insured = int(primeassurance.prime) / insured_count
            Insured.objects.filter(responsible=primeassurance.responsible).update(prime=prime_per_insured)
    
    else:
        insured.delete()
        insured_count = Insured.objects.filter(responsible=primeassurance.responsible).count()
        if insured_count > 0:
            prime_per_insured = int(primeassurance.prime) / insured_count
            Insured.objects.filter(responsible=primeassurance.responsible).update(prime=prime_per_insured)

    return redirect('view_insured')

@permission_required('amana.view_provider', raise_exception=True)
def view_provider(request):
    search_query = request.POST.get('search', '')
    provider = Provider.objects.all()
    if search_query:
        provider = provider.filter(Q(responsable__icontains=search_query) | Q(name__icontains=search_query) | Q(type__icontains=search_query))
    provider = provider.order_by('-id')
    paginator = Paginator(provider, 12) 
    page_number = request.GET.get('page', 1) 
    page = paginator.get_page(page_number)
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
       'providers':page,
       'search_query':search_query
    }
    return render(request, 'provider/view_provider.html', context)

@permission_required('amana.add_provider', raise_exception=True)
def add_provider(request):
    if request.method == 'POST':
        responsable = request.POST.get('responsable')
        type = request.POST.get('type')
        name = request.POST.get('name')
        provider = Provider.objects.create(
            responsable = responsable,
            type=type,
            name=name
        )
        provider.save()
        return redirect('view_provider')
    current_user = request.user
    first_name_current = current_user.first_name.capitalize()
    last_name_current = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name_current[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'first_name': first_name_current,
        'last_name': last_name_current,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
    }
    return render(request, 'provider/add_provider.html', context)

@permission_required('amana.change_provider', raise_exception=True)
def change_provider(request, provider_id):
    provider = get_object_or_404(Provider, id=provider_id)
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'provider': provider,
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
    }
    
    if request.method == 'POST':
        responsable = request.POST.get('responsable')
        type = request.POST.get('type')
        name = request.POST.get('name')
        
        provider.name = name
        provider.type = type
        provider.responsable = responsable
        provider.save()
        
        return redirect('view_provider')
    
    return render(request, 'provider/change_provider.html', context)

@permission_required('amana.delete_provider', raise_exception=True)
def delete_provider(request, provider_id):
    provider = get_object_or_404(Provider, id=provider_id)
    provider.delete()
    return redirect('view_provider')

@permission_required('amana.view_invoicehealth', raise_exception=True)
def view_invoicehealth(request):
    search_query = request.POST.get('search', '')
    invoice = InvoiceHealth.objects.all()
    if search_query:
        invoice = invoice.filter(Q(invoice_number__icontains=search_query) | Q(registration_number__icontains=search_query) | Q(responsable__icontains=search_query))
    invoice = invoice.order_by('-date')
    paginator = Paginator(invoice, 12) 
    page_number = request.GET.get('page', 1) 
    page = paginator.get_page(page_number)
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
       'invoices':page,
       'search_query':search_query
    }
    return render(request, 'invoicehealth/view_invoicehealth.html', context)

@permission_required('amana.view_invoicehealth', raise_exception=True)
def invoicehealth(request):
    search_query = request.POST.get('search', '')
    invoice = InvoiceHealth.objects.all()
    if search_query:
        invoice = invoice.filter(Q(provider__icontains=search_query) | Q(description__icontains=search_query) | Q(invoice_number__icontains=search_query) | Q(registration_number__icontains=search_query) | Q(responsible__icontains=search_query))
    invoice = invoice.order_by('-date')
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': current_user.is_superuser,
       'invoices':invoice,
       'search_query':search_query
    }
    return render(request, 'invoicehealth/view.html', context)

@permission_required('amana.add_invoicehealth', raise_exception=True)
def add_invoicehealth(request):
    if request.method == 'POST':
        try:
            invoices_data = json.loads(request.POST['data'])
            invoice_objects = [
                InvoiceHealth(
                    date=invoice['date'],
                    agency=invoice['agency'],
                    provider=invoice['provider'],
                    insured=invoice['insured'],
                    month=invoice['month'],
                    responsible=invoice['responsible'],
                    invoice_number=invoice['invoice_number'],
                    registration_number=invoice['registration_number'],
                    description=invoice['description'],
                    amount_th=invoice['amount_th']
                )
                for invoice in invoices_data
            ]
            InvoiceHealth.objects.bulk_create(invoice_objects)
            return redirect('view_invoicehealth')
        except (KeyError, json.JSONDecodeError) as e:
            # Handle error (optional)
            pass
    
    registration_numbers = Insured.objects.values('registration_number').distinct()
    providers = Provider.objects.all()
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'providers': providers,
        'registration_numbers': registration_numbers,
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': request.user.is_superuser,
    }
    return render(request, 'invoicehealth/add_invoicehealth.html', context)

@permission_required('amana.change_invoicehealth', raise_exception=True)
def change_invoicehealth(request,invoicehealth_id):
    invoicehealth = get_object_or_404(InvoiceHealth, id=invoicehealth_id)
    if request.method == 'POST':
        date = request.POST.get('date')
        invoice_number = request.POST.get('invoice_number')
        registration_number = request.POST.get('registration_number')
        agency = request.POST.get('agency')
        month = request.POST.get('month')
        responsible = request.POST.get('responsible')
        description = request.POST.get('description')
        insured = request.POST.get('insured')
        provider = request.POST.get('provider')
        amount_th = request.POST.get('amount_th')
        invoicehealth.date = date
        invoicehealth.invoice_number = invoice_number
        invoicehealth.registration_number = registration_number
        invoicehealth.agency = agency
        invoicehealth.month = month
        invoicehealth.responsible = responsible
        invoicehealth.description = description
        invoicehealth.insured = insured
        invoicehealth.provider = provider
        invoicehealth.amount_th = amount_th
        invoicehealth.save()
        return redirect('view_invoicehealth')
    registration_numbers = Insured.objects.values('registration_number').distinct()
    providers = Provider.objects.all()
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'
    context = {
        'providers': providers,
        'registration_numbers': registration_numbers,
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': request.user.is_superuser,
        'invoicehealth':invoicehealth
    }
    return render(request, 'invoicehealth/change_invoicehealth.html', context)

@permission_required('amana.delete_invoicehealth', raise_exception=True)
def delete_invoicehealth(request, invoicehealth_id):
    invoicehealth = get_object_or_404(InvoiceHealth, id=invoicehealth_id)
    invoicehealth.delete()
    return redirect('view_invoicehealth')

def get_insured_info(request):
    registration_number = request.GET.get('registration_number')
    insured_data = Insured.objects.filter(registration_number=registration_number).values('insured')
    responsable = Insured.objects.filter(registration_number=registration_number).values_list('responsible', flat=True).first()

    insured_list = list(insured_data)

    data = {
        'insured': insured_list,
        'responsable': responsable
    }
    return JsonResponse(data)


def view_summary(request):
    context = {}
    if request.method == 'POST':
        responsible = request.POST.get('responsible')
        month = request.POST.get('month')
        year = request.POST.get('year')

        if not responsible or not month or not year:
            context['error'] = 'Tous les champs sont obligatoires.'
            return render(request, 'summary/summary.html', context)

        invoices = InvoiceHealth.objects.filter(
            responsible=responsible,
            date__year=year
        )
        if month != "Année":
            invoices = invoices.filter(month=month)

        if not invoices.exists():
            context['error'] = 'Aucune facture trouvée pour les critères sélectionnés.'
            return render(request, 'summary/summary.html', context)
        
        primeassurance = PrimeAssurance.objects.filter(
        responsible=responsible
        )
        if primeassurance.exists():
            # Prendre le premier objet PrimeAssurance trouvé
            prime_assurance_obj = primeassurance.first()
            delta = relativedelta(prime_assurance_obj.date_fin, prime_assurance_obj.date_emission)
        else:
            context['error'] = 'Aucune assurance trouvée pour le responsable sélectionné.'
            return render(request, 'summary/summary.html', context)

        total_consumption = invoices.aggregate(Sum('amount_th'))['amount_th__sum'] or Decimal(0)

        if month == "Année":
            current_month = datetime.now().month
            insured_queryset = Insured.objects.filter(
                registration_number__in=invoices.values('registration_number')
            ).annotate(
                expected_consumption=ExpressionWrapper(
                    (F('prime') / Decimal(delta.years * 12 + delta.months)) * Decimal(current_month),
                    output_field=DecimalField()
                )
            )
        else:
            insured_queryset = Insured.objects.filter(
                registration_number__in=invoices.values('registration_number')
            ).annotate(
                expected_consumption=ExpressionWrapper(
                    F('prime') / Decimal(delta.years * 12 + delta.months),
                    output_field=DecimalField()
                )
            )

        insured_dict = {insured.registration_number: insured.expected_consumption for insured in insured_queryset}


        consumption_data = invoices.values(
            'registration_number',
            'insured'
        ).annotate(
            total_consumed=Sum('amount_th')
        ).order_by('-total_consumed')

        file_path = os.path.join(settings.BASE_DIR, 'amana/summary.xlsx') 
        try:
            workbook = load_workbook(file_path)
        except FileNotFoundError:
            context['error'] = f'Le fichier Excel {file_path} est introuvable.'
            return render(request, 'summary/summary.html', context)

        sheet = workbook.active
        start_row = 5
        end_row = 9
        current_row = start_row

        border = Border(
            left=Side(border_style="thin", color="999999"),
            right=Side(border_style="thin", color="999999"),
            top=Side(border_style="thin", color="999999"),
            bottom=Side(border_style="thin", color="999999")
        )

        def set_cell(sheet, cell, value, border):
            sheet[cell] = value
            sheet[cell].border = border
            sheet[cell].number_format = '#,##0.00'

        for data in consumption_data:
            if current_row > end_row:
                sheet.insert_rows(end_row + 1)
                end_row += 1

            registration_number = data['registration_number']
            expected_consumption = insured_dict.get(registration_number, Decimal(0))

            set_cell(sheet, f'A{current_row}', registration_number, border)
            set_cell(sheet, f'B{current_row}', data['insured'], border)
            set_cell(sheet, f'C{current_row}', data['total_consumed'], border)
            set_cell(sheet, f'D{current_row}', expected_consumption, border)

            current_row += 1

        total_row = end_row + 1
        total_expected_consumption = Decimal(0)
        for row_num in range(start_row, total_row):
            cell_value = sheet[f'D{row_num}'].value
            if cell_value:
                total_expected_consumption += Decimal(cell_value)
                
        set_cell(sheet, f'A{total_row}', "Total", border)
        set_cell(sheet, f'C{total_row}', total_consumption, border)
        set_cell(sheet, f'D{total_row}', total_expected_consumption, border)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f'rapport_summary_{responsible}_{month}_{year}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)

        return response

    primeassurance = PrimeAssurance.objects.all()
    current_user = request.user
    first_name = current_user.first_name.capitalize()
    last_name = current_user.last_name.capitalize()
    first_letter_of_first_name = first_name[0].upper()
    user_sex = SexUser.objects.filter(user=current_user).first()
    sex_name = user_sex.sex.sex_name.capitalize() if user_sex else 'Not specified'

    context.update({
        'primeassurance': primeassurance,
        'first_name': first_name,
        'last_name': last_name,
        'first_letter_of_first_name': first_letter_of_first_name,
        'sex_name': sex_name,
        'superuser': request.user.is_superuser,
    })

    return render(request, 'summary/summary.html', context)
